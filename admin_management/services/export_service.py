"""
Service for exporting data to CSV and Excel formats.
Supports exporting orders, payments, users, and financial reports.
"""
import csv
import io
from typing import List, Dict, Any, Optional
from datetime import datetime
from decimal import Decimal

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExportService:
    """Service for exporting data to various formats."""
    
    @staticmethod
    def export_to_csv(data: List[Dict[str, Any]], filename: str = 'export.csv') -> tuple:
        """
        Export data to CSV format.
        
        Args:
            data: List of dictionaries containing the data to export
            filename: Name for the exported file
            
        Returns:
            Tuple of (HttpResponse, filename)
        """
        from django.http import HttpResponse
        
        if not data:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response, filename
        
        # Get all unique keys from all dictionaries
        fieldnames = set()
        for row in data:
            fieldnames.update(row.keys())
        fieldnames = sorted(list(fieldnames))
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.DictWriter(response, fieldnames=fieldnames)
        writer.writeheader()
        
        for row in data:
            # Convert values to strings, handling None and special types
            cleaned_row = {}
            for key, value in row.items():
                if value is None:
                    cleaned_row[key] = ''
                elif isinstance(value, (datetime,)):
                    cleaned_row[key] = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, Decimal):
                    cleaned_row[key] = str(value)
                elif isinstance(value, (dict, list)):
                    cleaned_row[key] = str(value)
                else:
                    cleaned_row[key] = value
            writer.writerow(cleaned_row)
        
        return response, filename
    
    @staticmethod
    def export_to_excel(data: List[Dict[str, Any]], filename: str = 'export.xlsx', sheet_name: str = 'Data') -> tuple:
        """
        Export data to Excel format.
        
        Args:
            data: List of dictionaries containing the data to export
            filename: Name for the exported file
            sheet_name: Name for the Excel sheet
            
        Returns:
            Tuple of (HttpResponse, filename)
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")
        
        from django.http import HttpResponse
        
        if not data:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response, filename
        
        # Get all unique keys from all dictionaries
        fieldnames = set()
        for row in data:
            fieldnames.update(row.keys())
        fieldnames = sorted(list(fieldnames))
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Header row styling
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Write headers
        for col_num, fieldname in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_num, value=fieldname)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data rows
        for row_num, row_data in enumerate(data, 2):
            for col_num, fieldname in enumerate(fieldnames, 1):
                value = row_data.get(fieldname, '')
                
                # Convert special types
                if isinstance(value, (datetime,)):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, Decimal):
                    value = float(value)
                elif isinstance(value, (dict, list)):
                    value = str(value)
                elif value is None:
                    value = ''
                
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        for col_num, fieldname in enumerate(fieldnames, 1):
            column_letter = get_column_letter(col_num)
            max_length = max(
                len(str(fieldname)),
                max([len(str(row.get(fieldname, ''))) for row in data] + [0])
            )
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        return response, filename
    
    @staticmethod
    def prepare_orders_export(orders) -> List[Dict[str, Any]]:
        """Prepare order data for export."""
        export_data = []
        for order in orders.select_related('client', 'assigned_writer', 'paper_type', 'academic_level', 'formatting_style', 'website'):
            export_data.append({
                'Order ID': order.id,
                'Topic': order.topic or '',
                'Client ID': order.client.id if order.client else '',
                'Client Email': order.client.email if order.client else '',
                'Writer ID': order.assigned_writer.id if order.assigned_writer else '',
                'Writer Email': order.assigned_writer.email if order.assigned_writer else '',
                'Paper Type': order.paper_type.name if order.paper_type else '',
                'Academic Level': order.academic_level.name if order.academic_level else '',
                'Formatting Style': order.formatting_style.name if order.formatting_style else '',
                'Number of Pages': order.number_of_pages or 0,
                'Number of Slides': order.number_of_slides or 0,
                'Status': order.status,
                'Total Price': float(order.total_price) if order.total_price else 0,
                'Writer Compensation': float(order.writer_compensation) if order.writer_compensation else 0,
                'Is Paid': 'Yes' if order.is_paid else 'No',
                'Client Deadline': order.client_deadline.strftime('%Y-%m-%d %H:%M:%S') if order.client_deadline else '',
                'Writer Deadline': order.writer_deadline.strftime('%Y-%m-%d %H:%M:%S') if order.writer_deadline else '',
                'Created At': order.created_at.strftime('%Y-%m-%d %H:%M:%S') if order.created_at else '',
                'Updated At': order.updated_at.strftime('%Y-%m-%d %H:%M:%S') if order.updated_at else '',
                'Website': order.website.name if order.website else '',
            })
        return export_data
    
    @staticmethod
    def prepare_payments_export(payments) -> List[Dict[str, Any]]:
        """Prepare payment data for export."""
        export_data = []
        for payment in payments.select_related('client', 'order', 'website'):
            export_data.append({
                'Payment ID': payment.id,
                'Reference ID': payment.reference_id or '',
                'Transaction ID': payment.transaction_id or '',
                'Client ID': payment.client.id if payment.client else '',
                'Client Email': payment.client.email if payment.client else '',
                'Order ID': payment.order.id if payment.order else '',
                'Payment Type': payment.payment_type,
                'Amount': float(payment.amount) if payment.amount else 0,
                'Original Amount': float(payment.original_amount) if payment.original_amount else 0,
                'Discounted Amount': float(payment.discounted_amount) if payment.discounted_amount else 0,
                'Status': payment.status,
                'Payment Method': payment.payment_method or '',
                'Created At': payment.created_at.strftime('%Y-%m-%d %H:%M:%S') if payment.created_at else '',
                'Confirmed At': payment.confirmed_at.strftime('%Y-%m-%d %H:%M:%S') if payment.confirmed_at else '',
                'Website': payment.website.name if payment.website else '',
            })
        return export_data
    
    @staticmethod
    def prepare_users_export(users) -> List[Dict[str, Any]]:
        """Prepare user data for export."""
        export_data = []
        for user in users.select_related('client_profile', 'writer_profile'):
            export_data.append({
                'User ID': user.id,
                'Username': user.username or '',
                'Email': user.email or '',
                'Role': user.role or '',
                'First Name': user.first_name or '',
                'Last Name': user.last_name or '',
                'Is Active': 'Yes' if user.is_active else 'No',
                'Is Staff': 'Yes' if user.is_staff else 'No',
                'Is Suspended': 'Yes' if getattr(user, 'is_suspended', False) else 'No',
                'Date Joined': user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else '',
                'Last Login': user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else '',
                'Registration ID (Client)': getattr(user.client_profile, 'registration_id', '') if hasattr(user, 'client_profile') and user.client_profile else '',
                'Registration ID (Writer)': getattr(user.writer_profile, 'registration_id', '') if hasattr(user, 'writer_profile') and user.writer_profile else '',
                'Pen Name': getattr(user.writer_profile, 'pen_name', '') if hasattr(user, 'writer_profile') and user.writer_profile else '',
            })
        return export_data
    
    @staticmethod
    def prepare_financial_report_export(financial_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Prepare financial overview data for export."""
        export_data = []
        
        # Summary section
        export_data.append({
            'Category': 'Summary',
            'Item': 'Total Revenue',
            'Amount': financial_data.get('total_revenue', 0),
            'Date': datetime.now().strftime('%Y-%m-%d')
        })
        export_data.append({
            'Category': 'Summary',
            'Item': 'Total Writer Payments',
            'Amount': financial_data.get('total_writer_payments', 0),
            'Date': datetime.now().strftime('%Y-%m-%d')
        })
        export_data.append({
            'Category': 'Summary',
            'Item': 'Total Expenses',
            'Amount': financial_data.get('total_expenses', 0),
            'Date': datetime.now().strftime('%Y-%m-%d')
        })
        export_data.append({
            'Category': 'Summary',
            'Item': 'Net Profit',
            'Amount': financial_data.get('net_profit', 0),
            'Date': datetime.now().strftime('%Y-%m-%d')
        })
        
        # Revenue by source
        revenue_by_source = financial_data.get('revenue_by_source', {})
        for source, amount in revenue_by_source.items():
            export_data.append({
                'Category': 'Revenue by Source',
                'Item': source,
                'Amount': amount,
                'Date': datetime.now().strftime('%Y-%m-%d')
            })
        
        return export_data

